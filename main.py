
import json
from openpyxl import load_workbook

# Carrega um arquivo Excel e lê uma planilha específica
arquivo = 'planilhaQuestoes.xlsx'
planilha = 'Planilha1'
wb = load_workbook(filename=arquivo, read_only=True, data_only=True)
ws = wb[planilha]

# Listas para armazenar dados
evento = []
line = 1
data = []

def isResponse(ws, line, value):
    """
    Verifica se uma célula é uma resposta correta.

    Args:
        ws (Worksheet): Planilha Excel.
        line (int): Número da linha.
        value (int): Valor da célula.

    Returns:
        int: 0 se não for resposta correta, o valor fornecido caso contrário.
    """
    fill = ws.cell(line, 1).fill
    if fill is not None and fill.start_color is not None:
        color_in_hex = ws.cell(line, 1).fill.start_color.index
        if color_in_hex == "FF00B050":
            return value

    return 0

def concatena_valores_celulas(ws, line):
    """
    Concatena os valores de duas células.

    Args:
        ws (Worksheet): Planilha Excel.
        line (int): Número da linha.

    Returns:
        str: String concatenada ou uma string vazia se alguma célula for nula.
    """
    cell_value_1 = ws.cell(line, 1).value
    cell_value_2 = ws.cell(line, 2).value

    return f"{cell_value_1} {cell_value_2}" if cell_value_1 is not None and cell_value_2 is not None else ""

def get_cell_value(ws, line, col):
    """
    Obtém o valor de uma célula.

    Args:
        ws (Worksheet): Planilha Excel.
        line (int): Número da linha.
        col (int): Número da coluna.

    Returns:
        str: Valor da célula ou uma string vazia se a célula for nula.
    """
    value = ws.cell(line, col).value
    return "" if value is None else value

def process_line(ws, line):
    """
    Processa uma linha da planilha.

    Args:
        ws (Worksheet): Planilha Excel.
        line (int): Número da linha.

    Returns:
        int: Próxima linha a ser processada.
    """
    evento.append(get_cell_value(ws, line, 1))

    if evento[-1] is not None:
        questao = concatena_valores_celulas(ws, line)
        line += 1
        resp = 0

        alternativas = []
        for i in range(4):
            print(i)
            if resp == 0:
                resp = isResponse(ws, line, i)
            
            alternativa = concatena_valores_celulas(ws, line)
            alternativas.append(alternativa)
            line += 1

        data.append({
            'pergunta': questao,
            'alternativas': alternativas,
            'resposta': resp
        })
        return line

try:
    while line <= 171:
        print(f"process_line {line}")
        line = process_line(ws, line)
        line += 1

except Exception as e:
    print(f"Erro na linha {line}: {str(e)}")

# Converte dados para JSON e salva em um arquivo
json_data = json.dumps(data, ensure_ascii=False, indent=4)

json_output_file = "output.json"
with open(json_output_file, 'w', encoding='utf-8') as f:
    f.write(json_data)
